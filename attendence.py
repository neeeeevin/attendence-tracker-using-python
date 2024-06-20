import openpyxl 
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 

# loading the excel sheet 
book = openpyxl.load_workbook(r'C:\Users\NEVIN\Desktop\attendance.xlsx') 

# Choose the sheet 
sheet = book['Sheet1'] 

# counting number of rows / students 
r = sheet.max_row 

# counting number of columns / subjects 
c = sheet.max_column 

# list of students to remind 
l1 = [] 

# to concatenate list of roll numbers with 
# lack of attendance 
l2 = "" 

# list of roll numbers with lack of attendance 
l3 = [] 

# staff mail ids 
staff_mails = ['neeeeevinn848@gmail.com', 'kevinmon848@gmail.com'] 

# Warning messages 
m1 = "Warning!!! You can take only one more day leave for dsa class."
m2 = "Warning!!! You can take only one more day leave for Python class."
m3 = "Warning!!! You can take only one more day leave for IoT class."

def savefile(): 
    book.save(r'C:\Users\NEVIN\Desktop\attendance.xlsx') 
    print("Saved!") 

def check(no_of_days, row_num, b): 
    global l2 
    global l3 

    for student in range(len(row_num)): 
        # if total no.of.leaves equals threshold 
        if no_of_days[student] == 2: 
            if b == 1: 
                l1.append(sheet.cell(row=row_num[student], column=2).value) 
                mailstu(l1, m1) # sending mail 
            elif b == 2: 
                l1.append(sheet.cell(row=row_num[student], column=2).value) 
                mailstu(l1, m2) 
            else: 
                l1.append(sheet.cell(row=row_num[student], column=2).value) 
                mailstu(l1, m3) 
        # if total.no.of.leaves > threshold 
        elif no_of_days[student] > 2: 
            l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
            l3.append(sheet.cell(row=row_num[student], column=2).value) 
            subject = ["DSA", "Python", "IoT"][b-1]

    # If threshold crossed, modify the message 
    if l2 and l3: 
        msg1 = "You have lack of attendance in " + subject + "!!!"
        msg2 = "The following students have lack of attendance in your subject: " + l2
        mailstu(l3, msg1) # mail to students 
        staff_id = staff_mails[b-1] # pick respective staff's mail_id 
        mailstaff(staff_id, msg2) # mail to staff 

# for students 
def mailstu(li, msg): 
    from_id = 'studentemail@gmail.com'
    pwd = 'trygivingpass'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120) 
    s.starttls() 
    s.login(from_id, pwd) 

    for to_id in li: 
        message = MIMEMultipart() 
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain')) 
        s.sendmail(from_id, to_id, message.as_string()) 
    s.quit() 
    print("Mail sent to students") 

# for staff 
def mailstaff(mail_id, msg): 
    from_id = 'alofyr321@gmail.com'
    pwd = 'password'
    message = MIMEMultipart() 
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain')) 
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120) 
    s.starttls() 
    s.login(from_id, pwd) 
    s.sendmail(from_id, mail_id, message.as_string()) 
    s.quit() 
    print('Mail Sent to staff') 

resp = 1
while resp == 1: 
    print("1--->dsa\n2--->Python\n3--->IoT") 

    # enter the corresponding number 
    y = int(input("Enter subject: ")) 

    # no.of.absentees for that subject 
    no_of_absentees = int(input('Number of absentees: ')) 

    if no_of_absentees > 1: 
        x = list(map(int, input('Roll nos: ').split())) 
    else: 
        x = [int(input('Roll no: '))] 

    row_num = [] 
    no_of_days = [] 

    for student in x: 
        for i in range(2, r + 1): 
            if sheet.cell(row=i, column=1).value == student: 
                if y == 1: 
                    m = sheet.cell(row=i, column=3).value 
                    sheet.cell(row=i, column=3).value = m + 1
                elif y == 2: 
                    m = sheet.cell(row=i, column=4).value 
                    sheet.cell(row=i, column=4).value = m + 1
                elif y == 3: 
                    m = sheet.cell(row=i, column=5).value 
                    sheet.cell(row=i, column=5).value = m + 1
                no_of_days.append(m + 1) 
                row_num.append(i) 
                savefile()

    check(no_of_days, row_num, y) 
    resp = int(input('Another subject? 1---->Yes 0--->No: ')) 
